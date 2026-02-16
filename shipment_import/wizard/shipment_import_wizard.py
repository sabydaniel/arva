from odoo import models, fields

class ShipmentImportWizard(models.TransientModel):
    _name = 'shipment.import.wizard'
    _description = 'Shipment Import Wizard'

    file = fields.Binary(string="Upload XLSX", required=True)
    file_name = fields.Char(string="File Name")

    def action_import(self):

        import base64
        import io
        import openpyxl

        file_data = base64.b64decode(self.file)
        workbook = openpyxl.load_workbook(io.BytesIO(file_data))
        sheet = workbook.active

        shipment_date = sheet['B2'].value
        customer = sheet['B3'].value

        shipment = self.env['shipment.shipment'].create({
            'name': customer or 'Shipment',
            'shipment_date': shipment_date,
            'customer': customer,
            'sales_person': sheet['B4'].value,
            'origin': sheet['E3'].value,
            'destination': sheet['E4'].value,
            'description': sheet['B5'].value,
        })

        start_row = None

        for row in range(1, sheet.max_row + 1):
            if sheet[f'A{row}'].value and "TRANSPORTATION" in str(sheet[f'A{row}'].value):
                start_row = row
                break

        if not start_row:
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'shipment.shipment',
                'view_mode': 'form',
                'res_id': shipment.id,
                'target': 'current',
            }

        row = start_row

        while row <= sheet.max_row:

            desc = sheet[f'A{row}'].value

            if not desc or desc == "Total":
                break

            buying_total = sheet[f'F{row}'].value
            selling_total = sheet[f'K{row}'].value

            cost = buying_total if isinstance(buying_total, (int, float)) else 0.0
            sell = selling_total if isinstance(selling_total, (int, float)) else 0.0

            self.env['shipment.charge'].create({
                'shipment_id': shipment.id,
                'description': desc,
                'cost': cost,
                'sell': sell,
            })

            row += 1

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'shipment.shipment',
            'view_mode': 'form',
            'res_id': shipment.id,
            'target': 'current',
        }